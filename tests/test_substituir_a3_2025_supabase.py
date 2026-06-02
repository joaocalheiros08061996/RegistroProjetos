from argparse import Namespace
from datetime import datetime
import json

from openpyxl import Workbook

from automacoes.substituir_a3_2025_supabase import (
    DEFAULT_WORKBOOK,
    PROCESS_CLASSIFICATION_EXISTING,
    build_records_2025,
    project_type_2025,
    read_excel_records_2025,
    run_replacement,
    time_entries_ready,
)


def test_default_workbook_uses_2025_file_without_revision_suffix():
    assert DEFAULT_WORKBOOK.name == "A3 - Gerenciamento de Projetos (2025).xlsx"


def _project_row(**overrides):
    row = {
        "PROJETOS": "Projeto A",
        "TIPO DE PROJETOS": "Melhoria Contínua dos Processos",
        "RESPONSÁVEL": "Fagner",
        "FTes": "#REF!",
        "DATA INÍCIO PLANEJ": datetime(2025, 1, 10),
        "DATA FIM PLANEJ": datetime(2025, 1, 8),
        "GRAVIDADE": "3- Grave",
        "URGÊNCIA": "3 - Urgente",
        "TENDêNCIA": "3 - Piora em médio prazo",
        "OBJETIVOS": "1 - Objetivo totalmente definido",
        "MÉTODOS": "2 - Métodos conhecidos com pequenas adaptações",
    }
    row.update(overrides)
    return row


def _task_row(**overrides):
    row = {
        "PROJETO": "Projeto A",
        "TAREFA": "Classificação",
        "SUBTAREFA": "Estudos",
        "RESPONSÁVEL": "Fagner",
        "DATA INÍCIO PLANEJ": datetime(2025, 1, 10),
        "DATA FIM DO PLANEJ": datetime(2025, 1, 11),
        "DATA INÍCIO REAL": datetime(2025, 1, 10),
        "DIAS REAIS": 1,
        "PERCENTUAL REALIZADO": 1,
    }
    row.update(overrides)
    return row


def test_legacy_project_type_is_imported_as_existing_improvement():
    project_type, process_classification = project_type_2025(
        "Melhoria Contínua dos Processos"
    )

    assert project_type == "MELHORIA"
    assert process_classification == PROCESS_CLASSIFICATION_EXISTING


def test_build_records_2025_skips_ref_rows_and_uses_compound_task_names():
    records = build_records_2025(
        project_rows=[
            _project_row(_row_number=2, **{"RESPONSÁVEL": "fagner"}),
            _project_row(
                _row_number=3,
                PROJETOS="Projeto quebrado",
                **{"DATA INÍCIO PLANEJ": "#REF!"},
            ),
        ],
        task_rows=[
            _task_row(_row_number=2),
            _task_row(
                _row_number=3,
                SUBTAREFA="Relatório",
                **{"DIAS REAIS": "0,5", "PERCENTUAL REALIZADO": 0.5},
            ),
            _task_row(_row_number=4),
            _task_row(_row_number=5, **{"DATA INÍCIO PLANEJ": "#REF!"}),
            _task_row(_row_number=6, PROJETO="Projeto quebrado"),
        ],
        year=2025,
    )

    assert len(records.projects) == 1
    project = records.projects[0]
    assert project.project_type == "MELHORIA"
    assert project.process_classification == PROCESS_CLASSIFICATION_EXISTING
    assert project.responsible_login == "Fagner"
    assert project.fte == 1.0
    assert project.planned_start == datetime(2025, 1, 10)
    assert project.planned_end == datetime(2025, 1, 10)

    assert [task.name for task in records.tasks] == [
        "Classificação - Estudos",
        "Classificação - Relatório",
    ]
    assert records.tasks[0].status == "COMPLETED"
    assert records.tasks[0].actual_seconds == 86400
    assert records.tasks[1].status == "PAUSED"
    assert records.tasks[1].actual_seconds == 43200
    assert time_entries_ready(records) == 2

    assert records.skipped_projects == [
        {"row": 3, "name": "Projeto quebrado", "reason": "ref_error"}
    ]
    assert [item["reason"] for item in records.skipped_tasks] == [
        "duplicate_task_name",
        "ref_error",
        "project_not_imported",
    ]


def test_read_excel_records_2025_reads_new_sheets(tmp_path):
    workbook = Workbook()
    projects = workbook.active
    projects.title = "PROJETOS_2025"
    projects.append(
        [
            "PROJETOS",
            "TIPO DE PROJETOS ",
            "RESPONSÁVEL",
            "FTes",
            "DATA INÍCIO PLANEJ",
            "DATA FIM PLANEJ",
            "GRAVIDADE",
            "URGÊNCIA",
            "TENDêNCIA",
            "OBJETIVOS",
            "MÉTODOS ",
        ]
    )
    projects.append(
        [
            "Projeto Excel",
            "Normatização dos Processos",
            "João Calheiros",
            1,
            datetime(2025, 2, 12),
            datetime(2025, 2, 19),
            "4 - Muito grave",
            "3 - Urgente",
            "5 - Piora rapidamente",
            "3 - Objetivo parcialmente definido",
            "4 - Métodos pouco definidos",
        ]
    )

    tasks = workbook.create_sheet("TAREFAS_2025")
    tasks.append(
        [
            "PROJETO",
            "TAREFA",
            "SUBTAREFA",
            "RESPONSÁVEL",
            "DATA INÍCIO PLANEJ",
            "DATA FIM DO PLANEJ",
            "DATA INÍCIO REAL",
            "DIAS REAIS",
            "PERCENTUAL REALIZADO",
        ]
    )
    tasks.append(
        [
            "Projeto Excel",
            "Classificação",
            "Estudos",
            "João Calheiros",
            datetime(2025, 2, 12),
            datetime(2025, 2, 13),
            datetime(2025, 2, 12),
            1,
            1,
        ]
    )

    workbook_path = tmp_path / "a3_2025.xlsx"
    workbook.save(workbook_path)

    records = read_excel_records_2025(workbook_path, year=2025)

    assert [project.name for project in records.projects] == ["Projeto Excel"]
    assert records.projects[0].project_type == "NORMATIZACAO"
    assert [task.name for task in records.tasks] == ["Classificação - Estudos"]


def test_run_replacement_dry_run_writes_report_without_database(monkeypatch, tmp_path):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    workbook = Workbook()
    projects = workbook.active
    projects.title = "PROJETOS_2025"
    projects.append(
        [
            "PROJETOS",
            "TIPO DE PROJETOS",
            "RESPONSÁVEL",
            "FTes",
            "DATA INÍCIO PLANEJ",
            "DATA FIM PLANEJ",
        ]
    )
    projects.append(
        [
            "Projeto Dry Run",
            "Exportação",
            "Fagner",
            1,
            datetime(2025, 7, 28),
            datetime(2025, 7, 31),
        ]
    )

    tasks = workbook.create_sheet("TAREFAS_2025")
    tasks.append(
        [
            "PROJETO",
            "TAREFA",
            "SUBTAREFA",
            "DATA INÍCIO PLANEJ",
            "DATA FIM DO PLANEJ",
            "DATA INÍCIO REAL",
            "DIAS REAIS",
            "PERCENTUAL REALIZADO",
        ]
    )
    tasks.append(
        [
            "Projeto Dry Run",
            "Simulação",
            "NA",
            datetime(2025, 7, 28),
            datetime(2025, 7, 31),
            datetime(2025, 7, 28),
            3,
            1,
        ]
    )

    workbook_path = tmp_path / "dry-run.xlsx"
    report_path = tmp_path / "report.json"
    workbook.save(workbook_path)

    report = run_replacement(
        Namespace(
            workbook=workbook_path,
            user_id="user-123",
            year=2025,
            report=report_path,
            commit=False,
        )
    )

    assert report["mode"] == "dry-run"
    assert report["counts"]["projects_ready"] == 1
    assert report["counts"]["tasks_ready"] == 1
    assert report["counts"]["time_entries_ready"] == 1
    assert report["counts"]["delete_candidates_checked"] is False
    assert json.loads(report_path.read_text(encoding="utf-8"))["user_id"] == "user-123"
